import io
import os
import re
import gc
import traceback
from datetime import datetime
import pandas as pd
import numpy as np
from pathlib import Path


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS DE LEITURA ROBUSTA  (baseados em CMDB_Unificar_Bases.py)
# ─────────────────────────────────────────────────────────────────────────────


def _fix_unbalanced_quotes(raw_text: str) -> str:
    """Corrige quebras de linha com aspas desbalanceadas em CSVs mal-formados."""
    lines = raw_text.replace("\r\n", "\n").replace("\r", "\n").split("\n")
    fixed_lines, buffer = [], ""
    for line in lines:
        buffer += ("\n" + line) if buffer else line
        if buffer.count('"') % 2 == 0:
            fixed_lines.append(buffer)
            buffer = ""
    if buffer:
        fixed_lines.append(buffer)
    return "\n".join(fixed_lines)


def _read_csv_robust(path) -> pd.DataFrame:
    """
    Le CSV com multiplas heuristicas de fallback.
    Ordem: utf-8 auto-sep -> latin1 auto-sep -> fix-quotes fallback.
    """
    path = Path(path)
    for enc in ("utf-8", "latin1"):
        try:
            return pd.read_csv(
                path,
                engine="python",
                dtype=str,
                sep=None,
                encoding=enc,
                on_bad_lines="skip",
            )
        except Exception:
            pass
    try:
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            raw = f.read()
        fixed = _fix_unbalanced_quotes(raw)
        return pd.read_csv(io.StringIO(fixed), engine="python", dtype=str, sep=None)
    except Exception:
        pass
    raise RuntimeError(f"Falha ao ler CSV apos todos os fallbacks: {path}")


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS DO ALEX_VISAO  (logica de nulos por classe)
# ─────────────────────────────────────────────────────────────────────────────


def _contar_nulos_por_id(df: pd.DataFrame, id_col: str, colunas) -> pd.DataFrame:
    """
    Transforma colunas em linhas contando nulos e total por id_col.
    Replica def contar_nulos_por_id do Alex_visao_base_CMDB_V4.PY.
    """
    resultados = []
    for col in colunas:
        agrupado = (
            df.groupby(id_col)[col]
            .agg(nulos=lambda x: x.isna().sum(), total=lambda x: len(x))
            .reset_index()
        )
        agrupado["coluna"] = col
        resultados.append(agrupado[[id_col, "coluna", "nulos", "total"]])
    return pd.concat(resultados, ignore_index=True)


def _contar_computer(
    df: pd.DataFrame, id_col: str, colunas, col_ref: str = "sys_class_name"
) -> pd.DataFrame:
    """
    Variante para a class 'Computer': agrupa por u_category dentro
    do subconjunto onde col_ref == 'Computer'.
    Replica def contar_computer do Alex_visao_base_CMDB_V4.PY.
    """
    resultados = []
    df_filtrado = df[df[col_ref] == "Computer"]
    for col in colunas:
        agrupado = (
            df_filtrado.groupby(id_col)[col]
            .agg(nulos=lambda x: x.isna().sum(), total=lambda x: len(x))
            .reset_index()
        )
        agrupado["coluna"] = col
        resultados.append(agrupado[[id_col, "coluna", "nulos", "total"]])
    return pd.concat(resultados, ignore_index=True)


# ─────────────────────────────────────────────────────────────────────────────
# HELPER: LE O MASTER CORRETAMENTE
# ─────────────────────────────────────────────────────────────────────────────


def _carregar_master_attributes(caminho_master: str) -> pd.DataFrame:
    """
    Le a aba 'Attributes' do Master, seleciona colunas 4:26,
    filtra Mandatory == 'Mandatory'.

    Estrutura validada no Modelo_de_Dados_-_Master_v4.2.xlsx:
      Colunas 4:26 = Module, Class, Sys Class Name, Level, Path, Attribute,
                     Variable, Type, Reference, Max length, Default value,
                     Definition, Section, Mandatory, Discovery, Automation,
                     Integration, Order
    Nota: NAO usa skiprows (o codigo de referencia usava uma versao antiga do Excel).
    """
    df = pd.read_excel(caminho_master, sheet_name="Attributes", engine="openpyxl")
    df = df.iloc[:, 4:26]
    df = df[df["Mandatory"] == "Mandatory"].copy()
    return df


# ─────────────────────────────────────────────────────────────────────────────
# DATA PROCESSOR
# ─────────────────────────────────────────────────────────────────────────────


class DataProcessor:

    def __init__(self):
        pass

    # ---------- utilitarios --------------------------------------------------

    def contar_linhas(self, path):
        try:
            with open(path, "r", encoding="utf-8", errors="ignore") as f:
                count = sum(1 for _ in f)
            return max(0, count - 1)
        except Exception:
            return 0

    def _tenta_ler_csv(self, filepath, nrows=None):
        tentativas = [
            {"enc": "utf-8", "sep": ";"},
            {"enc": "latin-1", "sep": ";"},
            {"enc": "utf-8-sig", "sep": ";"},
            {"enc": "utf-8", "sep": ","},
            {"enc": "latin-1", "sep": ","},
        ]
        last_exc = None
        for t in tentativas:
            try:
                df = pd.read_csv(
                    filepath,
                    encoding=t["enc"],
                    sep=t["sep"],
                    nrows=nrows,
                    on_bad_lines="skip",
                    engine="python",
                )
                if df.shape[1] > 1:
                    return df
            except Exception as e:
                last_exc = e
        if last_exc:
            raise last_exc
        raise ValueError("Formato CSV nao reconhecido.")

    # ---------- historico ----------------------------------------------------

    def listar_historico(self, root_path, dias_filtro):
        dados = []
        try:
            limite_data = None
            if dias_filtro and isinstance(dias_filtro, int):
                from datetime import timedelta

                limite_data = datetime.now() - timedelta(days=dias_filtro)
            for root, dirs, files in os.walk(root_path):
                pasta_atual = os.path.basename(root)
                is_mes = re.match(r"^\d{2}\.\d{4}$", pasta_atual)
                for file in files:
                    if file.lower().endswith(".csv") and "master" not in file.lower():
                        full_path = os.path.join(root, file)
                        try:
                            stats = os.stat(full_path)
                            dt_mod = datetime.fromtimestamp(stats.st_mtime)
                            if limite_data and dt_mod < limite_data:
                                continue
                            tag = "old" if (datetime.now() - dt_mod).days > 7 else "new"
                            nome_relatorio = os.path.basename(os.path.dirname(root))
                            if is_mes:
                                grupo_visual = pasta_atual
                                nome_visual = file
                            else:
                                grupo_visual = f"Outros > {pasta_atual}"
                                nome_visual = file
                            qtd_linhas = self.contar_linhas(full_path)
                            dados.append(
                                {
                                    "nome": nome_visual,
                                    "linhas": qtd_linhas,
                                    "data": dt_mod.strftime("%d/%m/%Y %H:%M"),
                                    "tag": tag,
                                    "caminho": full_path,
                                    "pasta_pai": grupo_visual,
                                    "origem_relatorio": nome_relatorio,
                                }
                            )
                        except Exception as e:
                            print(f"Aviso: '{file}': {e}")
        except Exception as e:
            print(f"Erro scan historico: {e}")
        dados.sort(
            key=lambda x: (
                x["pasta_pai"],
                datetime.strptime(x["data"], "%d/%m/%Y %H:%M"),
            ),
            reverse=True,
        )
        return dados, len(dados)

    def verificar_pendencias(self, root_path, mes_ano_alvo, lista_esperada_nomes):
        arquivos_na_pasta = set()
        pastas_encontradas = []
        for root, dirs, files in os.walk(root_path):
            nome_pasta = os.path.basename(root)
            if nome_pasta == mes_ano_alvo:
                pastas_encontradas.append(root)
                for f in files:
                    if f.lower().endswith(".csv"):
                        nome_sem_ext = os.path.splitext(f)[0]
                        nome_limpo = re.sub(
                            r"_Copia_\d+", "", nome_sem_ext, flags=re.IGNORECASE
                        )
                        arquivos_na_pasta.add(nome_limpo.lower())
                        arquivos_na_pasta.add(nome_sem_ext.lower())
        caminho_display = (
            f"{len(pastas_encontradas)} pastas '{mes_ano_alvo}' encontradas."
        )
        if not pastas_encontradas:
            return (
                lista_esperada_nomes,
                "Nenhuma pasta com este mes encontrada na rede.",
            )
        faltantes = [
            n for n in lista_esperada_nomes if n.lower() not in arquivos_na_pasta
        ]
        return faltantes, caminho_display

    def verificar_integridade(self, path):
        bad_files = []
        for root, _, files in os.walk(path):
            for f in files:
                if f.lower().endswith(".csv"):
                    full = os.path.join(root, f)
                    try:
                        self._tenta_ler_csv(full, nrows=5)
                    except Exception as e:
                        bad_files.append((f, str(e)))
        return bad_files

    # ---------- gerar master (CMDB_Unificar_Bases.py) ------------------------

    def gerar_master(self, root_path, mes_alvo, output_path=None):
        """
        Unifica todos os CSVs/Excel do mes em um unico arquivo MASTER.

        Logica primaria (CMDB_Unificar_Bases.py):
          ROOT_DIR / ClasseDir / MesDir / arquivos
          Adiciona coluna 'Module' = nome da ClasseDir.
          Saida padrao: ROOT_DIR / '1. Base Completa' / CMDB_{mes_alvo}.csv

        Fallback: varredura recursiva para qualquer pasta cujo nome == mes_alvo.
        """
        print(f"\nGerador Master | Mes: {mes_alvo} | Raiz: {root_path}")

        root = Path(root_path)
        rows = []
        arquivos_ok = 0
        arquivos_erro = []

        # abordagem primaria: ROOT/Modulo/Mes/arquivos
        for classe_dir in sorted(root.iterdir()):
            if not classe_dir.is_dir():
                continue
            month_folder = classe_dir / mes_alvo
            if not month_folder.exists():
                continue
            modulo = classe_dir.name
            print(f"  Modulo: {modulo}")
            for file in sorted(month_folder.glob("*.*")):
                if file.suffix.lower() not in (".csv", ".xlsx", ".xls"):
                    continue
                if "master" in file.name.lower() or file.name.startswith("~$"):
                    continue
                try:
                    if file.suffix.lower() == ".csv":
                        df = _read_csv_robust(file)
                    else:
                        df = pd.read_excel(
                            file, sheet_name=0, dtype=str, engine="openpyxl"
                        )
                    df["Module"] = modulo
                    rows.append(df)
                    arquivos_ok += 1
                    print(f"    OK  {file.name}  ({len(df):,} linhas)")
                except Exception as e:
                    arquivos_erro.append(f"{file.name}: {e}")
                    print(f"    ERRO  {file.name}: {e}")

        # fallback: varredura recursiva
        if not rows:
            print("  Abordagem primaria sem resultados — varredura recursiva...")
            for r, _dirs, files in os.walk(root_path):
                if os.path.basename(r) == mes_alvo:
                    modulo = os.path.basename(os.path.dirname(r))
                    print(f"  Pasta encontrada: {r}")
                    for fname in sorted(files):
                        ext = os.path.splitext(fname)[1].lower()
                        if ext not in (".csv", ".xlsx", ".xls"):
                            continue
                        if "master" in fname.lower() or fname.startswith("~$"):
                            continue
                        full = Path(r) / fname
                        try:
                            df = (
                                _read_csv_robust(full)
                                if ext == ".csv"
                                else pd.read_excel(full, sheet_name=0, dtype=str)
                            )
                            df["Module"] = modulo
                            rows.append(df)
                            arquivos_ok += 1
                            print(f"    OK  {fname}  ({len(df):,} linhas)")
                        except Exception as e:
                            arquivos_erro.append(f"{fname}: {e}")

        if not rows:
            print(f"  Nenhum arquivo encontrado para o mes {mes_alvo}.")
            return None

        try:
            df_final = pd.concat(rows, ignore_index=True, sort=False)
            cols_ordered = ["Module"] + [c for c in df_final.columns if c != "Module"]
            df_final = df_final.reindex(columns=cols_ordered)

            if not output_path:
                nome_arq = f"CMDB_{mes_alvo}.csv"
                output_dir = root / "1. Base Completa"
                output_dir.mkdir(parents=True, exist_ok=True)
                output_path = str(output_dir / nome_arq)

            if os.path.exists(output_path):
                os.remove(output_path)

            df_final.to_csv(output_path, index=False, encoding="utf-8-sig")
            print(f"\n  Master gerado: {output_path}")
            print(f"  Arquivos processados: {arquivos_ok}")
            print(f"  Total de linhas:      {len(df_final):,}")
            if arquivos_erro:
                print(f"  Erros ignorados:      {len(arquivos_erro)}")
            return output_path

        except Exception as e:
            print(f"Erro critico ao consolidar: {e}")
            return None

    # ---------- unificar partes ----------------------------------------------

    def unificar_partes(self, pasta_alvo, output_path=None):
        if (
            not pasta_alvo
            or not os.path.exists(pasta_alvo)
            or not os.path.isdir(pasta_alvo)
        ):
            return {
                "sucesso": False,
                "msg": f"Diretorio invalido: {pasta_alvo}",
                "unificados": 0,
                "erros": [],
            }

        grupos = {}
        for f in os.listdir(pasta_alvo):
            if f.lower().endswith(".csv") and re.search(
                r"_pt\d+", f, flags=re.IGNORECASE
            ):
                base_name = re.sub(r"_pt\d+", "", f, flags=re.IGNORECASE)
                grupos.setdefault(base_name, []).append(os.path.join(pasta_alvo, f))

        if not grupos:
            return {
                "sucesso": True,
                "msg": "Nenhum arquivo com partes (_pt1, _pt2...) encontrado.",
                "unificados": 0,
                "erros": [],
            }

        unificados = 0
        erros = []
        for base, arquivos in grupos.items():
            if len(arquivos) < 2:
                continue
            dfs = []
            for arq in sorted(arquivos):
                try:
                    dfs.append(self._tenta_ler_csv(arq))
                except Exception as e:
                    erros.append(f"Leitura de '{os.path.basename(arq)}': {e}")
            if not dfs:
                erros.append(f"Nenhuma parte legivel para '{base}'")
                continue
            try:
                df_final = pd.concat(dfs, ignore_index=True)
                nome_final = f"UNIFICADO_{base}"
                if not nome_final.lower().endswith(".csv"):
                    nome_final += ".csv"
                df_final.to_csv(
                    os.path.join(pasta_alvo, nome_final),
                    sep=";",
                    index=False,
                    encoding="utf-8-sig",
                )
                unificados += 1
            except Exception as e:
                erros.append(f"Concat/escrita de '{base}': {e}")

        msg = f"{unificados} grupo(s) unificado(s) com sucesso."
        if erros:
            msg += f" | {len(erros)} erro(s)."
        return {"sucesso": True, "msg": msg, "unificados": unificados, "erros": erros}

    # --- LÓGICA DO RELATÓRIO DE COMPLETUDE (ATUALIZADA) ---
    def gerar_relatorio_completude(self, caminho_cmdb, caminho_master, caminho_saida):
        """Lógica avançada para cruzar a Base CMDB com as regras de completude do Master."""
        try:
            # 1. CARREGAMENTO E MAPEAMENTO DO MASTER
            try:
                # Busca dinâmica da aba para evitar erro de case-sensitive ou espaços
                xl = pd.ExcelFile(caminho_master)
                aba_alvo = next(
                    (
                        s
                        for s in xl.sheet_names
                        if s.strip().lower() in ["attributes", "atributos"]
                    ),
                    None,
                )

                if not aba_alvo:
                    return (
                        False,
                        f"Aba 'Attributes' não encontrada no Master. Abas disponíveis: {xl.sheet_names}",
                    )

                df_master_raw = pd.read_excel(
                    caminho_master, sheet_name=aba_alvo, header=None
                )
                header_idx = -1
                for idx, row in df_master_raw.head(30).iterrows():
                    row_str = row.astype(str).str.lower().tolist()
                    if "variable" in row_str and (
                        "class" in row_str or "sys class name" in row_str
                    ):
                        header_idx = idx
                        break

                df_master = pd.read_excel(
                    caminho_master,
                    sheet_name=aba_alvo,
                    skiprows=header_idx if header_idx != -1 else 0,
                )
            except Exception as e:
                return False, f"Erro ao ler arquivo Master: {str(e)}"

            df_master.columns = df_master.columns.astype(str).str.strip().str.lower()

            # CLONAGEM DE MÚLTIPLAS COLUNAS NO MASTER
            mapa_clonagem = {
                "company": ["company", "u_company"],
                "business_criticality": [
                    "business_criticality",
                    "u_business_criticality",
                ],
                "sys_class_name": ["sys_class_name", "class", "sys class name"],
                "u_type_ref": ["u_type_ref", "type_ref"],
            }

            for destino, origens in mapa_clonagem.items():
                col_nome_filtro = f"{destino}.filtro"
                col_origem_master = next(
                    (c for c in df_master.columns if c in origens), None
                )
                if col_origem_master:
                    df_master[col_nome_filtro] = df_master[col_origem_master].copy()

            col_mandatoria = next(
                (c for c in df_master.columns if "mandatory" in c), None
            )
            col_atributo_master = (
                "variable" if "variable" in df_master.columns else None
            )
            candidatos_classe = [
                c
                for c in df_master.columns
                if c in ["class", "sys class name", "sys_class_name"]
            ]

            if not col_mandatoria or not col_atributo_master or not candidatos_classe:
                return False, f"Erro: Colunas cruciais não encontradas no Master."

            master_filtrado = df_master[
                df_master[col_mandatoria]
                .astype(str)
                .str.contains("Mandatory", case=False, na=False)
            ].copy()
            vars_mandatorias = set(
                master_filtrado[col_atributo_master].astype(str).str.strip().unique()
            )

            # 2. ANÁLISE DO CSV
            df_header = self._tenta_ler_csv(caminho_cmdb, nrows=0)
            cols_csv_orig = [c.strip() for c in df_header.columns.astype(str)]

            mapa_colunas_csv = {}
            col_classe_real = next(
                (c for c in cols_csv_orig if c.lower() in ["sys_class_name", "class"]),
                None,
            )
            if not col_classe_real:
                return False, "A coluna 'sys_class_name' não foi encontrada no CSV."
            mapa_colunas_csv["sys_class_name"] = col_classe_real

            col_company_real = next(
                (
                    c
                    for c in cols_csv_orig
                    if c.lower() in ["company", "u_company", "company_name"]
                ),
                None,
            )
            if col_company_real:
                mapa_colunas_csv["company"] = col_company_real

            col_bus_crit_real = next(
                (
                    c
                    for c in cols_csv_orig
                    if c.lower() in ["business_criticality", "u_business_criticality"]
                ),
                None,
            )
            if col_bus_crit_real:
                mapa_colunas_csv["business_criticality"] = col_bus_crit_real

            col_type_ref_real = next(
                (c for c in cols_csv_orig if c.lower() in ["u_type_ref", "type_ref"]),
                None,
            )
            if col_type_ref_real:
                mapa_colunas_csv["u_type_ref"] = col_type_ref_real

            # 3. DETECÇÃO DE MATCH
            df_amostra = self._tenta_ler_csv(caminho_cmdb, nrows=5000)
            classes_csv = set(
                df_amostra[col_classe_real]
                .dropna()
                .astype(str)
                .str.strip()
                .str.upper()
                .unique()
            )

            melhor_coluna_master = None
            maior_match = 0
            for col_cand in candidatos_classe:
                classes_master = set(
                    master_filtrado[col_cand]
                    .dropna()
                    .astype(str)
                    .str.strip()
                    .str.upper()
                    .unique()
                )
                match_count = len(classes_csv.intersection(classes_master))
                if match_count > maior_match:
                    maior_match = match_count
                    melhor_coluna_master = col_cand

            if not melhor_coluna_master:
                melhor_coluna_master = candidatos_classe[0]

            # 4. CARREGAMENTO E CLONAGEM NO CSV
            cols_base_csv = list(mapa_colunas_csv.values())
            cols_para_carregar = [
                c for c in cols_csv_orig if c in cols_base_csv or c in vars_mandatorias
            ]

            df_fonte = pd.read_csv(
                caminho_cmdb,
                sep=None,
                engine="python",
                encoding="utf-8-sig",
                usecols=cols_para_carregar,
                dtype=str,
                on_bad_lines="skip",
            )
            df_fonte.columns = df_fonte.columns.astype(str).str.strip()

            colunas_agrupamento_finais = []
            chaves_ordem = [
                "sys_class_name",
                "company",
                "business_criticality",
                "u_type_ref",
            ]

            for chave in chaves_ordem:
                nome_col_filtro = f"{chave}.filtro"
                if chave in mapa_colunas_csv:
                    nome_col_real = mapa_colunas_csv[chave]
                    if nome_col_real in df_fonte.columns:
                        df_fonte[nome_col_filtro] = df_fonte[nome_col_real].copy()
                        colunas_agrupamento_finais.append(nome_col_filtro)

            valores_nulos = [
                "",
                "No Data",
                "Without Data",
                "N/A",
                "null",
                "nan",
                "None",
            ]
            df_fonte.replace(valores_nulos, np.nan, inplace=True)
            gc.collect()

            grouped = df_fonte.groupby(colunas_agrupamento_finais, dropna=False)
            df_totais = grouped.size().reset_index(name="Total")
            df_preenchidos = grouped.count().reset_index()

            del df_fonte
            gc.collect()

            df_melted = df_preenchidos.melt(
                id_vars=colunas_agrupamento_finais,
                var_name="Atributo",
                value_name="Preenchidos",
            )
            df_calculo = pd.merge(
                df_melted, df_totais, on=colunas_agrupamento_finais, how="left"
            )

            # 5. O JOIN FINAL
            cols_master_select = [melhor_coluna_master, col_atributo_master]
            master_join = master_filtrado[cols_master_select].rename(
                columns={
                    melhor_coluna_master: "Classe_Alvo",
                    col_atributo_master: "Atributo_Alvo",
                }
            )

            master_join["Classe_Alvo"] = (
                master_join["Classe_Alvo"].astype(str).str.strip().str.upper()
            )
            master_join["Atributo_Alvo"] = (
                master_join["Atributo_Alvo"].astype(str).str.strip().str.upper()
            )

            col_join_classe = (
                "sys_class_name.filtro"
                if "sys_class_name.filtro" in df_calculo.columns
                else colunas_agrupamento_finais[0]
            )
            df_calculo["Classe_Join"] = (
                df_calculo[col_join_classe].astype(str).str.strip().str.upper()
            )
            df_calculo["Atributo_Join"] = (
                df_calculo["Atributo"].astype(str).str.strip().str.upper()
            )

            df_final = pd.merge(
                df_calculo,
                master_join,
                left_on=["Classe_Join", "Atributo_Join"],
                right_on=["Classe_Alvo", "Atributo_Alvo"],
                how="inner",
            )

            if df_final.empty:
                return (
                    False,
                    f"Relatório 0 linhas. Verifique compatibilidade de Classes.",
                )

            df_final["Total"] = df_final["Total"].fillna(0).astype(int)
            df_final["Preenchidos"] = df_final["Preenchidos"].fillna(0).astype(int)

            # Ajuste de Nulos explicitamente na tabela de saída
            df_final["Nulos"] = df_final["Total"] - df_final["Preenchidos"]

            df_final["Percentual_Preenchido"] = np.where(
                df_final["Total"] > 0, df_final["Preenchidos"] / df_final["Total"], 0
            ).round(4)

            df_final["Percentual_Nulos"] = np.where(
                df_final["Total"] > 0, df_final["Nulos"] / df_final["Total"], 0
            ).round(4)

            df_final["ColunaExiste"] = True

            colunas_remover = [
                "Classe_Join",
                "Atributo_Join",
                "Classe_Alvo",
                "Atributo_Alvo",
            ]
            cols_saida = [c for c in df_final.columns if c not in colunas_remover]
            df_final[cols_saida].to_excel(caminho_saida, index=False)

            return True, f"Sucesso! Relatório gerado em: {caminho_saida}"
        except Exception as e:
            return False, f"Erro Técnico:\n{traceback.format_exc()}"


# ─────────────────────────────────────────────────────────────────────────────
# STRUCTURE VALIDATOR
# ─────────────────────────────────────────────────────────────────────────────


class StructureValidator:
    """
    Valida se os campos do Master estao presentes como
    colunas na base CMDB. Gera relatorio Excel com CHECK por campo/classe.
    """

    def __init__(self):
        self.master_data = None
        self.master_path = None
        self._data_processor = DataProcessor()

    def carregar_master(self, caminho_master, force_reload=False):
        if (
            self.master_data is not None
            and self.master_path == caminho_master
            and not force_reload
        ):
            return self.master_data
        try:
            self.master_path = caminho_master
            if caminho_master.lower().endswith((".xlsx", ".xls")):
                df = pd.read_excel(caminho_master, sheet_name=None, engine="openpyxl")
                self.master_data = df
                return df
            df = self._data_processor._tenta_ler_csv(caminho_master)
            self.master_data = {"Base": df}
            return self.master_data
        except Exception as e:
            self.master_data = None
            raise Exception(f"Falha ao carregar master '{caminho_master}': {e}")

    def executar_validacao_completa(
        self, caminho_cmdb: str, caminho_master: str, caminho_saida: str
    ):
        """
        Valida campos do Master vs colunas da base CMDB.
        Gera Excel com CHECK Presente/Ausente por campo e classe.
        """
        try:
            # 1. Master
            print("\n-- Carregando Master (executar_validacao_completa) --")
            try:
                # Carregamos diretamente para aplicar a regra de filtro (Mandatory, Optional, No)
                df_master_raw = pd.read_excel(
                    caminho_master, sheet_name="Attributes", engine="openpyxl"
                )
                df_mandatorios = df_master_raw.iloc[:, 4:26]

                # Filtra a coluna Mandatory para trazer as 3 opções
                valores_permitidos = ["MANDATORY", "OPTIONAL", "NO"]
                df_mandatorios = df_mandatorios[
                    df_mandatorios["Mandatory"]
                    .astype(str)
                    .str.strip()
                    .str.upper()
                    .isin(valores_permitidos)
                ].copy()
            except Exception as e:
                return False, (
                    f"Erro ao ler o arquivo Master:\n{e}\n\n"
                    "Verifique se o arquivo possui a aba 'Attributes'."
                )

            if df_mandatorios.empty:
                return False, (
                    "Nenhum campo com status válido (Mandatory, Optional, No) encontrado no Master.\n"
                    "Verifique a coluna 'Mandatory'."
                )
            print(f"   Campos totais validados: {len(df_mandatorios)}")
            print(f"   Classes:            {df_mandatorios['Class'].nunique()}")

            # 2. CMDB (Verificação Real de Dados e Ativos)
            print("\n-- Lendo e validando base do CMDB --")
            try:
                # O leitor robusto trará a base para podermos checar se há DADOS reais
                df_cmdb = self._data_processor._tenta_ler_csv(caminho_cmdb)
                cols_cmdb_orig = list(df_cmdb.columns)
                cols_cmdb_upper = set(str(c).strip().upper() for c in cols_cmdb_orig)

                col_classe_csv = next(
                    (
                        c
                        for c in cols_cmdb_orig
                        if str(c).strip().lower() in ["sys_class_name", "class"]
                    ),
                    None,
                )

                arquivo_vazio = df_cmdb.empty
                classes_com_ativos = set()

                if col_classe_csv and not arquivo_vazio:
                    # Mapeia quais classes realmente possuem registros dentro do arquivo
                    classes_com_ativos = set(
                        df_cmdb[col_classe_csv]
                        .dropna()
                        .astype(str)
                        .str.strip()
                        .str.upper()
                    )

                print(f"   Colunas no CMDB: {len(cols_cmdb_upper)}")
                print(f"   Total de ativos encontrados: {len(df_cmdb)}")

                # Libera a memória do dataframe do CMDB, pois já temos os cabeçalhos e as classes ativas
                del df_cmdb
                import gc

                gc.collect()

            except Exception as e:
                return False, f"Nao foi possivel ler o arquivo CMDB:\n{e}"

            # 3. Cruzamento campo a campo
            print("\n-- Cruzando campos do Master vs colunas CMDB --")
            resultados = []
            campos_vistos = set()

            col_module = "Module"
            col_attribute = "Attribute"
            col_classe = "Class"
            col_atributo = "Variable"
            col_descricao = (
                "Definition" if "Definition" in df_mandatorios.columns else None
            )
            col_mandatory = "Mandatory"

            for _, row in df_mandatorios.iterrows():
                atributo = str(row.get(col_atributo, "")).strip()
                if not atributo or atributo.upper() in ("NAN", "NONE", ""):
                    continue

                classe = str(row.get(col_classe, "")).strip()
                if not classe or pd.isna(row.get(col_classe)):
                    classe = "N/A"

                chave = (classe.upper(), atributo.upper())
                if chave in campos_vistos:
                    continue
                campos_vistos.add(chave)

                # Resgatando Module e Attribute
                modulo = str(row.get(col_module, "")).strip()
                if modulo.upper() in ("NAN", "NONE"):
                    modulo = ""

                atributo_desc = str(row.get(col_attribute, "")).strip()
                if atributo_desc.upper() in ("NAN", "NONE"):
                    atributo_desc = ""

                descricao = (
                    str(row.get(col_descricao, "")).strip()
                    if col_descricao and pd.notna(row.get(col_descricao))
                    else ""
                )

                # --- LÓGICA CORRIGIDA DOS VERDADEIROS / FALSOS ---
                existe_coluna = atributo.upper() in cols_cmdb_upper

                if arquivo_vazio:
                    existe = False
                    status_str = "Ausente (Base Vazia)"
                    observacao = "O arquivo não possui base de ativos (0 linhas)."
                elif (
                    col_classe_csv
                    and classe.upper() not in classes_com_ativos
                    and classe.upper() != "N/A"
                ):
                    existe = False
                    status_str = "Ausente (Classe sem Ativos)"
                    observacao = f"A classe '{classe}' não possui ativos nesta base."
                elif not existe_coluna:
                    existe = False
                    status_str = "Ausente no CMDB"
                    observacao = "Campo validado não existe como coluna na base CMDB."
                else:
                    existe = True
                    status_str = "Presente no CMDB"
                    observacao = "OK"

                # Inserindo na ordem solicitada
                resultados.append(
                    {
                        "Module": modulo,
                        "Attribute": atributo_desc,
                        "Classe (sys_class_name)": classe,
                        "Campo (Variable)": atributo,
                        "Mandatory": str(row.get(col_mandatory, "")).strip(),
                        "Definicao": descricao,
                        "CHECK": existe,
                        "Status": status_str,
                        "Observacao": observacao,
                    }
                )

            if not resultados:
                return False, "Nenhum resultado gerado. Verifique os arquivos."

            df_resultado = pd.DataFrame(resultados)

            # Ordenação do DataFrame final (agrupa por check, módulo, classe e campo)
            df_resultado = df_resultado.sort_values(
                ["CHECK", "Module", "Classe (sys_class_name)", "Campo (Variable)"]
            )

            presentes = int(df_resultado["CHECK"].sum())
            ausentes = len(df_resultado) - presentes
            print(f"   Presentes: {presentes}  |  Ausentes: {ausentes}")

            # 4. Salva Excel formatado
            print(f"\n-- Salvando em: {caminho_saida} --")
            with pd.ExcelWriter(caminho_saida, engine="openpyxl") as writer:
                df_resultado.to_excel(writer, index=False, sheet_name="Validacao CMDB")
                ws = writer.sheets["Validacao CMDB"]

                for col_cells in ws.columns:
                    max_len = max(
                        (len(str(c.value)) if c.value is not None else 0)
                        for c in col_cells
                    )
                    ws.column_dimensions[col_cells[0].column_letter].width = min(
                        max_len + 4, 70
                    )

                from openpyxl.styles import PatternFill, Font

                hdr_fill = PatternFill(
                    start_color="1F3864", end_color="1F3864", fill_type="solid"
                )
                hdr_font = Font(color="FFFFFF", bold=True)
                for cell in ws[1]:
                    cell.fill = hdr_fill
                    cell.font = hdr_font

                red_fill = PatternFill(
                    start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"
                )
                green_fill = PatternFill(
                    start_color="CCFFCC", end_color="CCFFCC", fill_type="solid"
                )
                check_col_idx = df_resultado.columns.get_loc("CHECK")
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    fill = green_fill if row[check_col_idx].value else red_fill
                    for cell in row:
                        cell.fill = fill

            return True, (
                f"Validacao concluida com sucesso!\n\n"
                f"Total de campos verificados: {len(df_resultado)}\n"
                f"Presentes na base CMDB:  {presentes}\n"
                f"Ausentes no CMDB:        {ausentes}\n\n"
                f"Relatorio salvo em:\n{caminho_saida}"
            )

        except Exception:
            return False, f"Erro tecnico na validacao:\n{traceback.format_exc()}"

    def validar_tabela(self, caminho_tabela, nome_aba_master=None, nome_tabela=None):
        try:
            if caminho_tabela.lower().endswith(".csv"):
                df_tabela = self._data_processor._tenta_ler_csv(caminho_tabela, nrows=5)
            else:
                df_tabela = pd.read_excel(caminho_tabela, nrows=5)
            if df_tabela is None or df_tabela.empty or len(df_tabela.columns) == 0:
                return False, "Tabela invalida ou vazia."
            colunas_tabela = set(df_tabela.columns.astype(str).str.strip().str.upper())
            if self.master_data is None:
                if not self.master_path:
                    self.master_path = self._encontrar_arquivo_master()
                    if not self.master_path:
                        return False, "ARQUIVO MASTER INACESSIVEL."
                self.carregar_master(self.master_path)
            modelo = self._encontrar_modelo_no_master(
                nome_aba_master, nome_tabela, colunas_tabela
            )
            if not modelo:
                return False, "Nenhum modelo compativel encontrado no Master."
            colunas_modelo = set(modelo["colunas"])
            faltantes = colunas_modelo - colunas_tabela
            extras = colunas_tabela - colunas_modelo
            correspondentes = colunas_tabela.intersection(colunas_modelo)
            percentual = (
                round(len(correspondentes) / len(colunas_modelo) * 100, 2)
                if colunas_modelo
                else 0
            )
            return True, {
                "tabela_analisada": caminho_tabela,
                "modelo_referencia": modelo["fonte"],
                "total_colunas_modelo": len(colunas_modelo),
                "total_colunas_tabela": len(colunas_tabela),
                "colunas_faltantes": sorted(faltantes),
                "colunas_extra": sorted(extras),
                "colunas_correspondentes": len(correspondentes),
                "conformidade_perfeita": not faltantes and not extras,
                "percentual_conformidade": percentual,
                "aba_master_encontrada": modelo["aba"],
                "similaridade": modelo["similaridade"],
            }
        except Exception as e:
            return False, f"Erro na validacao de '{caminho_tabela}': {e}"

    def _encontrar_arquivo_master(self):
        caminho_base = os.environ.get("AUDIT_MASTER_PATH", "")
        for nome in (
            "Modelo de Dados - Master_v4.2.xlsx",
            "Modelo de Dados - Master_v4.2.csv",
        ):
            c = os.path.join(caminho_base, nome) if caminho_base else ""
            if c and os.path.exists(c):
                return c
        return None

    def _encontrar_modelo_no_master(self, nome_aba, nome_tabela, colunas_tabela):
        melhor = None
        maior_sim = 0
        for aba, df in self.master_data.items():
            if len(df.columns) < 1:
                continue
            for colunas_master in self._extrair_colunas_master_candidatas(df):
                sim = self._calcular_similaridade(colunas_tabela, colunas_master)
                if sim > maior_sim:
                    maior_sim = sim
                    melhor = {
                        "fonte": f"Aba '{aba}'",
                        "aba": aba,
                        "colunas": colunas_master,
                        "similaridade": sim,
                    }
        return melhor if melhor and maior_sim > 0.05 else None

    def _extrair_colunas_master_candidatas(self, df_master):
        candidatas = []
        dfs = [df_master]
        for i in range(min(20, len(df_master))):
            vals = df_master.iloc[i].astype(str).str.lower()
            if any(
                t in v
                for v in vals.values
                for t in ["atributo", "variable", "coluna", "name", "class", "type"]
            ):
                df_des = df_master.iloc[i + 1 :].copy()
                df_des.columns = df_master.iloc[i].values
                dfs.append(df_des)
        for df in dfs:
            for col in df.columns:
                if any(
                    t in str(col).lower()
                    for t in ["atributo", "variable", "coluna", "field", "nome", "name"]
                ):
                    serie = df[col].dropna().astype(str).str.strip().str.upper()
                    colunas = set(serie[serie != ""])
                    if colunas:
                        candidatas.append(colunas)
            cab = set(df.columns.astype(str).str.strip().str.upper())
            validas = {
                c for c in cab if c and not c.startswith("UNNAMED") and c != "NAN"
            }
            if validas:
                candidatas.append(validas)
        return candidatas

    def _calcular_similaridade(self, colunas_tabela, colunas_master):
        if not colunas_tabela or not colunas_master:
            return 0
        return len(colunas_tabela.intersection(colunas_master)) / len(colunas_master)

    def definir_master_manual(self, caminho_master):
        self.master_path = caminho_master
        self.carregar_master(caminho_master, force_reload=True)


def validar_estrutura_tabela(caminho_tabela, caminho_master=None, nome_aba=None):
    """Funcao de conveniencia para validacao rapida."""
    v = StructureValidator()
    if caminho_master:
        v.definir_master_manual(caminho_master)
    return v.validar_tabela(caminho_tabela, nome_aba)
